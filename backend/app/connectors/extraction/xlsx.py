"""
Role: xlsx / xlsm (OOXML SpreadsheetML) attachment extraction (design §2.5) — openpyxl in
      read-only + data-only mode. A spreadsheet is STRUCTURED data, not prose, so this extractor
      produces BOTH a bounded text render (embeddable — the finding surface) AND a faithful TYPED
      cell grid (lossless, the source of truth, persisted to email_attachment.extracted_data for
      analysis-at-query-time later, e.g. via DuckDB). Dumb lossless ingest now; smart query later.
Used by: the IMAP connector's attachment_extractor (dispatches the two SpreadsheetML content types
         — .xlsx + macro-enabled .xlsm); scripts.backfill_attachment_extraction (via the seam).
         The arrow points IN — this module imports nothing back from any connector.
Depends on: openpyxl (MIT; et-xmlfile MIT underneath) — verified GPL-free; sibling extraction
            modules only: .extraction_result (the result contract incl. the `structured` field),
            .common (serialize_table — the ONE table serializer + MAX_EXTRACTED_CHARS + package_
            version), .ooxml (is_ole_container + validate_ooxml_package — the SHARED OOXML zip-bomb
            battery docx + xlsx both inherit), .text_sanitize (sanitize_body_text — the SINGLE
            stored-text sanitization source). Imports NOTHING from any specific connector.
Key invariants:
  - extract_xlsx_text NEVER raises: every failure (openpyxl / its zip layer / et-xmlfile raise
    diverse internals on malformed payloads) degrades to an ExtractionResult; a final catch-all
    guards even our own bugs.
  - PASSWORD-PROTECTED xlsx arrives as an OLE compound file (magic D0 CF 11 E0), not a zip
    (ECMA-376 encryption wraps the package) → `encrypted`. No password handling at MVP (§2.3).
  - The SHARED OOXML zip guards (EOCD pre-gate / member-count / decompressed-expansion /
    DOM-parsed-XML-parts bounds) run BEFORE openpyxl opens anything — a spreadsheet is a zip and
    inherits docx's bomb defenses verbatim (extraction.ooxml).
  - data_only=True yields the CACHED COMPUTED VALUES Excel last wrote — NOT formulas. This is a
    deliberate v1 limit: a library-generated workbook NEVER OPENED BY EXCEL has no cached values
    for its formula cells (they read back as None) and those cells are simply skipped. read_only=
    True streams rows so memory stays bounded on large sheets. Corpus measurement of cached-value
    availability is recorded in the slice's PR notes.
  - The TYPED grid (result.structured, format 'xlsx-grid-v1') is SPARSE (None/empty cells skipped)
    and BOUNDED: MAX_CELLS non-empty cells across the workbook AND a ~JSON-size backstop — over
    either, structured.truncated=true, capture STOPS, but everything captured so far is KEPT. At
    most MAX_SHEETS sheets are read. Cell types: number→'n', date/datetime→'d' (ISO-8601 string),
    bool→'b', everything else→'s' (string, sanitized). bool is checked BEFORE number (Python's
    bool is a subclass of int). number_format rides as 'nf' only when non-default ('General'). A
    NON-FINITE float (inf/-inf/nan — an overflowed cached formula value Excel last wrote) is NOT
    JSON/JSONB-storable, so it is captured as a STRING ('s', its str() form) rather than 'n': the
    numeric twin of sanitize_body_text's lone-surrogate strip, keeping the cell instead of letting
    the default json serializer emit the Infinity/NaN literals Postgres JSONB rejects.
  - The TEXT render is for embedding/finding ONLY — the structured grid is the source of truth.
    Per sheet a '[sheet: NAME]' marker + the captured cells as serialize_table rows, capped at
    RENDER_MAX_ROWS x RENDER_MAX_COLS so a wide/tall sheet can't explode the text ('[+more rows]'
    when row-capped); sheets joined by blank lines; MAX_EXTRACTED_CHARS applies (→ `truncated`).
  - Status: `extracted` (text rendered), `truncated` (TEXT over MAX_EXTRACTED_CHARS — a cell/size-
    bounded structured grid with full text stays `extracted` with structured.truncated=true),
    `empty` (no non-empty cell anywhere), `corrupt` (openpyxl/zip raised — class names only in
    detail, NEVER payload bytes), `encrypted`. detail (EQ-7) e.g. 'sheets=3 cells=1204' or
    'sheets=3 cells=50000 structured-truncated'. Provenance: extractor 'openpyxl' + its version.
  - `detail` carries exception CLASS NAMES / fixed phrases / counts only — never str(exc) and never
    cell content. openpyxl/et-xmlfile were LIVE-CHECKED to emit zero log records on parse failure
    and to keep payload bytes out of their exception strings (this module's leak regression test
    pins it) — no vendor-logger muting needed, unlike pypdf/pdfminer in pdf.py.
  - The caller (the seam) enforces the global size ceiling BEFORE this module parses anything.
"""

from __future__ import annotations

import logging
import math
from datetime import date, datetime, time
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

from app.connectors.extraction.common import (
    MAX_EXTRACTED_CHARS,
    package_version,
    serialize_table,
)
from app.connectors.extraction.extraction_result import (
    STATUS_CORRUPT,
    STATUS_EMPTY,
    STATUS_ENCRYPTED,
    STATUS_EXTRACTED,
    STATUS_TRUNCATED,
    ExtractionResult,
)
from app.connectors.extraction.ooxml import is_ole_container, validate_ooxml_package
from app.connectors.extraction.text_sanitize import sanitize_body_text

logger = logging.getLogger(__name__)

_OPENPYXL = "openpyxl"

# The structured-grid format tag — versioned so a later, smarter capture (e.g. one that resolves
# formulas, or stores merged-cell spans) can be told apart from this v1 cached-values grid.
_GRID_FORMAT = "xlsx-grid-v1"

# Typed-capture bounds (the structured grid is the lossless source of truth, but still bounded so a
# pathological workbook can't balloon a JSONB row / the analysis pipeline):
#   - MAX_CELLS: total NON-EMPTY cells captured across the WHOLE workbook.
#   - MAX_SCANNED_CELLS: total cell VISITS (empty included) across the workbook. read_only
#     iter_rows pads to the sheet's declared dimension, so a tiny attachment with one cell near
#     Excel's lower-right limit (XFD1048576, ~17e9 phantom cells) would otherwise spin the ingest
#     worker through billions of EMPTY cells without ever incrementing the non-empty MAX_CELLS
#     counter (2026-06-13 Codex review). This caps total work regardless of how sparse the sheet.
#   - MAX_STRUCTURED_BYTES: a JSON-size backstop (~8 MB) — checked WITH the candidate cell's size
#     BEFORE appending, so a single huge cell can't slip past (2026-06-13 Codex review).
#   - MAX_SHEETS: a workbook declaring more tabs than this is pathological; only the first are read.
# Hitting any of these sets structured.truncated=true and STOPS capture (kept-what-we-have).
MAX_CELLS = 50_000
MAX_SCANNED_CELLS = 5_000_000
MAX_STRUCTURED_BYTES = 8 * 1024 * 1024
MAX_SHEETS = 100

# Text-render caps (the render is the embeddable surface, NOT the source of truth) — a wide or tall
# sheet must not explode the stored text. Rows beyond the cap are dropped with a '[+more rows]'
# marker; columns beyond the cap are simply not rendered (the structured grid keeps them all).
RENDER_MAX_ROWS = 200
RENDER_MAX_COLS = 50

# OOXML number_format value meaning "no explicit format" — omitted from the typed cell ('nf' rides
# only when the producer set a real format, e.g. a currency or date mask).
_DEFAULT_NUMBER_FORMAT = "General"


def extract_xlsx_text(payload: bytes) -> ExtractionResult:
    """Extract an xlsx/xlsm workbook as BOTH a typed cell grid + a text render; never raises.

    Pipeline: OLE-magic probe (password-protected → `encrypted`) → the shared OOXML zip guards
    (EOCD/member/expansion bounds → `corrupt` before openpyxl opens anything) → openpyxl
    read-only + data-only load → typed sparse capture into result.structured ('xlsx-grid-v1',
    bounded by MAX_CELLS / MAX_STRUCTURED_BYTES / MAX_SHEETS) + a bounded per-sheet text render
    (serialize_table rows under a '[sheet: NAME]' marker) → sanitize → MAX_EXTRACTED_CHARS cap.

    Status: no non-empty cell anywhere → `empty` (structured still carries the empty sheet shells);
    text over the cap → `truncated`; a cell/size-bounded grid with full text → `extracted` with
    structured.truncated=true; an openpyxl/zip crash → `corrupt` (class name only in detail).

    Args:
        payload: the raw xlsx/xlsm bytes (the seam has already enforced the global size ceiling).

    Returns:
        An ExtractionResult; `text` is non-None for `extracted`/`truncated`, and `structured`
        carries the typed grid whenever openpyxl opened the workbook (None only on the
        encrypted/corrupt pre-parse exits).
    """
    try:
        return _extract(payload)
    except Exception as unexpected:  # the seam's NEVER-raise contract: even our bugs degrade
        # Class name ONLY - no exc_info: a formatted traceback ends with "Class: str(exc)", and
        # library exception strings could embed payload fragments (mirrors docx/pdf posture).
        logger.warning("xlsx extraction: unexpected failure (%s)", type(unexpected).__name__)
        return ExtractionResult(
            None, STATUS_CORRUPT, detail=f"unexpected:{type(unexpected).__name__}"
        )


def _extract(payload: bytes) -> ExtractionResult:
    """The real pipeline (see extract_xlsx_text); may raise — the public wrapper degrades."""
    if is_ole_container(payload):
        return ExtractionResult(
            None,
            STATUS_ENCRYPTED,
            detail="ole-container (password-protected xlsx)",
        )
    package_verdict = validate_ooxml_package(payload)
    if package_verdict is not None:
        return package_verdict
    try:
        workbook = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as load_error:  # BadZipFile / InvalidFileException / parse errors
        return ExtractionResult(
            None, STATUS_CORRUPT, detail=f"openpyxl:{type(load_error).__name__}"
        )
    try:
        capture = _capture_workbook(workbook)
    finally:
        workbook.close()  # read_only mode holds the zip open until closed
    return _result_from_capture(capture)


class _Capture:
    """The running typed grid + text render as the workbook is streamed (one pass, bounded)."""

    def __init__(self) -> None:
        self.sheets: list[dict] = []  # the structured 'sheets' list (xlsx-grid-v1)
        self.render_blocks: list[str] = []  # per-sheet text-render blocks
        self.cell_count = 0  # non-empty cells captured so far (across the workbook)
        self.scanned = 0  # cell VISITS so far, empty included (the phantom-cell DoS bound)
        self.structured_bytes = 0  # running JSON-ish size estimate (the byte backstop)
        self.truncated = False  # a cell/size/sheet/scan bound was hit — capture stopped early
        self.any_cell = False  # at least one non-empty cell anywhere (else → `empty`)


def _capture_workbook(workbook: openpyxl.Workbook) -> _Capture:
    """Stream each sheet (≤ MAX_SHEETS) into a typed sparse grid + a bounded text render.

    Stops capturing cells the moment a bound (MAX_CELLS / MAX_STRUCTURED_BYTES) is hit — but
    keeps every cell captured up to that point and records structured.truncated=true. Sheets past
    MAX_SHEETS are dropped (also a truncation). The text render is built from the SAME captured
    cells so the two views never disagree.
    """
    capture = _Capture()
    worksheets = workbook.worksheets
    if len(worksheets) > MAX_SHEETS:
        capture.truncated = True
        worksheets = worksheets[:MAX_SHEETS]
    for worksheet in worksheets:
        if capture.truncated and capture.cell_count >= MAX_CELLS:
            break  # already at the hard cell cap — no point opening more sheets
        _capture_sheet(worksheet, capture)
    return capture


def _capture_sheet(worksheet: openpyxl.worksheet.worksheet.Worksheet, capture: _Capture) -> None:
    """Capture one sheet's non-empty cells (typed) + build its text-render block.

    The typed cells are SPARSE (None/empty skipped). The render is a bounded 2D grid built from the
    captured cells (≤ RENDER_MAX_ROWS x RENDER_MAX_COLS) so a wide sheet can't explode the text.
    Dims report the sheet's declared extent (max_row/max_column), not the captured subset.
    """
    cells: list[dict] = []
    # render_grid[row_index] -> {col_index: text}; sparse, capped to the render window.
    render_grid: dict[int, dict[int, str]] = {}
    render_row_capped = False
    max_rendered_row = 0
    max_rendered_col = 0

    for row_index, row in enumerate(worksheet.iter_rows(), start=1):
        for column_index, cell in enumerate(row, start=1):
            capture.scanned += 1
            if capture.scanned > MAX_SCANNED_CELLS:
                # A sparse sheet with a bloated declared dimension would otherwise spin here
                # through billions of phantom empty cells (Codex review) — cap total visits.
                capture.truncated = True
                break
            typed = _typed_cell(cell, column_index, row_index)
            if typed is None:
                continue  # sparse: skip empty cells
            capture.any_cell = True
            cell_bytes = _estimate_cell_bytes(typed)
            # Check the bounds WITH this candidate cell, BEFORE appending: a single huge cell must
            # not slip past the byte backstop and leave truncated=false (Codex review).
            if (
                capture.cell_count >= MAX_CELLS
                or capture.structured_bytes + cell_bytes > MAX_STRUCTURED_BYTES
            ):
                capture.truncated = True
                break
            cells.append(typed)
            capture.cell_count += 1
            capture.structured_bytes += cell_bytes
            rendered, capped = _add_to_render(
                render_grid, row_index, column_index, typed["v"], cell.value
            )
            render_row_capped = render_row_capped or capped
            if rendered:
                max_rendered_row = max(max_rendered_row, row_index)
                max_rendered_col = max(max_rendered_col, column_index)
        else:
            continue  # inner loop completed without break — keep scanning rows
        # inner loop broke on a hard bound — stop this sheet too
        capture.truncated = True
        break

    capture.sheets.append(
        {
            "name": worksheet.title,
            "dims": {"rows": worksheet.max_row or 0, "cols": worksheet.max_column or 0},
            "cells": cells,
        }
    )
    capture.render_blocks.append(
        _render_block(
            worksheet.title,
            render_grid,
            max_rendered_row,
            max_rendered_col,
            render_row_capped,
        )
    )


def _typed_cell(cell: object, column_index: int, row_index: int) -> dict | None:
    """One non-empty cell as a typed dict {ref, t, v, [nf]}; None for an empty/whitespace cell.

    Types (bool BEFORE number — Python's bool is a subclass of int): bool→'b', number→'n'
    (int/float as a JSON number), date/datetime/time→'d' (ISO-8601 string), everything else→'s'
    (string, sanitized via sanitize_body_text so the stored value is UTF-8-safe). number_format
    rides as 'nf' only when the producer set a non-default ('General') format.

    NON-FINITE GUARD (the numeric twin of sanitize_body_text's lone-surrogate strip): a cached
    formula value that overflowed IEEE-754 surfaces under data_only as float('inf')/('-inf')/
    ('nan'). The default json serializer (database.py sets no override) emits these as the bare
    literals Infinity/-Infinity/NaN, which Postgres JSONB REJECTS — turning the very 'analysis-ready
    number' this slice exists to preserve into a poison message that dead-letters the whole email.
    So a non-finite number is captured as a STRING ('s', its str() form 'inf'/'-inf'/'nan') — the
    cell is preserved (not silently dropped), JSONB-safe, and honestly typed as non-numeric.
    """
    value = cell.value
    if value is None:
        return None
    if isinstance(value, bool):
        type_code, json_value = "b", value
    elif isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            # inf/-inf/nan is not JSONB-storable — keep it as its string form (see docstring).
            type_code, json_value = "s", str(value)
        else:
            type_code, json_value = "n", value
    elif isinstance(value, (datetime, date, time)):
        type_code, json_value = "d", value.isoformat()
    else:
        sanitized = sanitize_body_text(str(value))
        if not sanitized:
            return None  # whitespace-only string is not worth a cell (stays sparse)
        type_code, json_value = "s", sanitized

    reference = f"{get_column_letter(column_index)}{row_index}"
    typed: dict = {"ref": reference, "t": type_code, "v": json_value}
    number_format = getattr(cell, "number_format", None)
    if number_format and number_format != _DEFAULT_NUMBER_FORMAT:
        typed["nf"] = number_format
    return typed


def _estimate_cell_bytes(typed: dict) -> int:
    """A cheap upper-ish estimate of a typed cell's JSON footprint (the byte-backstop budget).

    Avoids json.dumps per cell (the hot loop runs MAX_CELLS times): the ref/type keys are a small
    fixed overhead, the value's str length dominates, and a present number_format adds its length.
    """
    overhead = 24  # {"ref":"","t":"","v":} punctuation + keys, approximately
    value_size = len(str(typed["v"]))
    format_size = len(typed["nf"]) + 8 if "nf" in typed else 0
    return overhead + len(typed["ref"]) + value_size + format_size


def _add_to_render(
    render_grid: dict[int, dict[int, str]],
    row_index: int,
    column_index: int,
    json_value: object,
    raw_value: object,
) -> tuple[bool, bool]:
    """Place a captured cell into the bounded render grid; returns (rendered?, row_capped?).

    Cells beyond RENDER_MAX_ROWS / RENDER_MAX_COLS are NOT rendered (the structured grid still has
    them). Dates render as their ISO string (the json_value), every other type as its str().
    """
    if column_index > RENDER_MAX_COLS:
        return False, False
    if row_index > RENDER_MAX_ROWS:
        return False, True
    text_value = json_value if isinstance(json_value, str) else str(raw_value)
    render_grid.setdefault(row_index, {})[column_index] = str(text_value)
    return True, False


def _render_block(
    sheet_name: str,
    render_grid: dict[int, dict[int, str]],
    max_row: int,
    max_col: int,
    row_capped: bool,
) -> str:
    """Build one sheet's text block: a '[sheet: NAME]' marker + serialize_table rows.

    Builds a dense 2D grid (empty cells → '') from the sparse render_grid up to the rendered
    extent, serializes it through the SHARED serialize_table (so xlsx rows look like every other
    extractor's tables), and appends '[+more rows]' when rows were dropped by the row cap. A sheet
    with no rendered cells contributes just its marker (its presence is still findable).
    """
    marker = f"[sheet: {sheet_name}]"
    if not render_grid:
        return marker
    grid: list[list[str | None]] = []
    for row_index in range(1, max_row + 1):
        row_cells = render_grid.get(row_index, {})
        grid.append([row_cells.get(column_index, "") for column_index in range(1, max_col + 1)])
    table_text = serialize_table(grid)
    block = f"{marker}\n{table_text}" if table_text else marker
    if row_capped:
        block = f"{block}\n[+more rows]"
    return block


def _result_from_capture(capture: _Capture) -> ExtractionResult:
    """Assemble the ExtractionResult from a finished capture: grid + text render + status.

    `empty` when no non-empty cell exists anywhere (text NULL, but the structured shells are kept
    so the workbook's shape is still recorded). Otherwise the joined render is sanitized and
    MAX_EXTRACTED_CHARS-capped (over the cap → `truncated`; a cell/size-bounded grid with full
    text stays `extracted` with structured.truncated=true). detail (EQ-7) reports sheet + cell
    counts, plus 'structured-truncated' when a typed bound fired.
    """
    structured = {
        "format": _GRID_FORMAT,
        "sheets": capture.sheets,
        "truncated": capture.truncated,
    }
    detail = f"sheets={len(capture.sheets)} cells={capture.cell_count}"
    if capture.truncated:
        detail = f"{detail} structured-truncated"
    version = package_version(_OPENPYXL)

    if not capture.any_cell:
        return ExtractionResult(
            None,
            STATUS_EMPTY,
            detail=detail,
            extractor_name=_OPENPYXL,
            extractor_version=version,
            structured=structured,
        )

    text = sanitize_body_text("\n\n".join(capture.render_blocks))
    if not text:
        # Cells existed but every render block sanitized to blank (e.g. all cells past the render
        # window): the typed grid is the truth, but there is no embeddable text — honest `empty`.
        return ExtractionResult(
            None,
            STATUS_EMPTY,
            detail=detail,
            extractor_name=_OPENPYXL,
            extractor_version=version,
            structured=structured,
        )
    if len(text) > MAX_EXTRACTED_CHARS:
        return ExtractionResult(
            text[:MAX_EXTRACTED_CHARS],
            STATUS_TRUNCATED,
            detail=f"{detail} text-capped from {len(text)} chars",
            extractor_name=_OPENPYXL,
            extractor_version=version,
            structured=structured,
        )
    return ExtractionResult(
        text,
        STATUS_EXTRACTED,
        detail=detail,
        extractor_name=_OPENPYXL,
        extractor_version=version,
        structured=structured,
    )
